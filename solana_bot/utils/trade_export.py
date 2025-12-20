"""
Trade History Export

Export trade history to CSV, JSON, and other formats for analysis.
"""

import csv
import json
import os
from datetime import datetime
from typing import List, Dict, Any, Optional
from dataclasses import dataclass, asdict
import logging

logger = logging.getLogger(__name__)


@dataclass
class TradeRecord:
    """Single trade record for export"""
    timestamp: str
    action: str  # BUY or SELL
    mint: str
    amount_sol: float
    token_amount: int
    signature: str
    dex: str
    reason: str  # For sells: TRAILING_STOP, TAKE_PROFIT, etc.
    pnl_sol: float
    pnl_pct: float
    hold_time_seconds: float
    success: bool
    error_message: str


class TradeExporter:
    """
    Export trade history to various formats.
    
    Supported formats:
    - CSV (spreadsheet compatible)
    - JSON (programmatic access)
    - Summary report (human readable)
    
    Usage:
        exporter = TradeExporter(db)
        exporter.export_csv("trades.csv")
        exporter.export_json("trades.json")
        print(exporter.generate_summary())
    """
    
    def __init__(self, db_manager):
        """
        Initialize exporter.
        
        Args:
            db_manager: DatabaseManager instance
        """
        self.db = db_manager
    
    def get_all_trades(self, limit: int = 1000) -> List[TradeRecord]:
        """Get all trades from database"""
        trades = []
        
        try:
            # Use DatabaseManager's get_trades method
            raw_trades = self.db.get_trades(limit=limit)
            
            for row in raw_trades:
                trades.append(TradeRecord(
                    timestamp=row.get('timestamp', ''),
                    action=row.get('action', ''),
                    mint=row.get('mint', ''),
                    amount_sol=row.get('amount_sol', 0.0) or 0.0,
                    token_amount=row.get('token_amount', 0) or 0,
                    signature=row.get('signature', '') or '',
                    dex=row.get('dex', '') or '',
                    reason="",  # Would need to parse from other data
                    pnl_sol=0.0,  # Calculated separately
                    pnl_pct=0.0,
                    hold_time_seconds=0.0,
                    success=bool(row.get('success', False)),
                    error_message=row.get('error_message', '') or ''
                ))
        
        except Exception as e:
            logger.error(f"Error fetching trades: {e}")
        
        return trades
    
    def export_csv(
        self,
        filepath: str,
        start_date: Optional[str] = None,
        end_date: Optional[str] = None
    ):
        """
        Export trades to CSV file.
        
        Args:
            filepath: Output file path
            start_date: Optional start date filter (YYYY-MM-DD)
            end_date: Optional end date filter (YYYY-MM-DD)
        """
        trades = self.get_all_trades()
        
        # Filter by date if specified
        if start_date or end_date:
            trades = self._filter_by_date(trades, start_date, end_date)
        
        if not trades:
            logger.warning("No trades to export")
            return
        
        with open(filepath, 'w', newline='', encoding='utf-8') as f:
            writer = csv.DictWriter(f, fieldnames=list(asdict(trades[0]).keys()))
            writer.writeheader()
            
            for trade in trades:
                writer.writerow(asdict(trade))
        
        logger.info(f"Exported {len(trades)} trades to {filepath}")
    
    def export_json(
        self,
        filepath: str,
        pretty: bool = True
    ):
        """
        Export trades to JSON file.
        
        Args:
            filepath: Output file path
            pretty: Pretty print JSON
        """
        trades = self.get_all_trades()
        
        data = {
            "export_date": datetime.now().isoformat(),
            "total_trades": len(trades),
            "trades": [asdict(t) for t in trades]
        }
        
        with open(filepath, 'w', encoding='utf-8') as f:
            if pretty:
                json.dump(data, f, indent=2)
            else:
                json.dump(data, f)
        
        logger.info(f"Exported {len(trades)} trades to {filepath}")
    
    def generate_summary(self) -> str:
        """Generate human-readable summary report"""
        trades = self.get_all_trades()
        
        if not trades:
            return "No trades found."
        
        # Calculate stats
        buys = [t for t in trades if t.action == "BUY"]
        sells = [t for t in trades if t.action == "SELL"]
        
        successful_buys = [t for t in buys if t.success]
        successful_sells = [t for t in sells if t.success]
        
        total_bought_sol = sum(t.amount_sol for t in successful_buys)
        
        # Get daily stats from database
        daily_stats = self._get_daily_stats()
        
        # Build report
        report = f"""
╔══════════════════════════════════════════════════════════════╗
║                    TRADE HISTORY SUMMARY                      ║
╠══════════════════════════════════════════════════════════════╣
║  Generated: {datetime.now().strftime("%Y-%m-%d %H:%M:%S"):^48} ║
╠══════════════════════════════════════════════════════════════╣
║  OVERALL STATISTICS                                          ║
╠══════════════════════════════════════════════════════════════╣
║  Total Trades:        {len(trades):>6}                                  ║
║  ├── Buys:            {len(buys):>6}  ({len(successful_buys)} successful)                ║
║  └── Sells:           {len(sells):>6}  ({len(successful_sells)} successful)                ║
║                                                              ║
║  Total SOL Invested:  {total_bought_sol:>10.4f} SOL                       ║
╠══════════════════════════════════════════════════════════════╣
║  DAILY BREAKDOWN                                             ║
╠══════════════════════════════════════════════════════════════╣
"""
        
        for stat in daily_stats[:7]:  # Last 7 days
            date = stat.get('date', 'N/A')
            total = stat.get('total_trades', 0)
            wins = stat.get('successful_trades', 0)
            win_rate = stat.get('win_rate', 0.0)
            pnl = stat.get('realized_pnl', 0.0)
            
            pnl_str = f"+{pnl:.4f}" if pnl >= 0 else f"{pnl:.4f}"
            
            report += f"║  {date}  │ {total:>3} trades │ {wins:>3} wins ({win_rate:>5.1f}%) │ {pnl_str:>10} SOL ║\n"
        
        report += """╚══════════════════════════════════════════════════════════════╝
"""
        
        return report
    
    def _get_daily_stats(self) -> List[Dict[str, Any]]:
        """Get daily statistics from database"""
        stats = []
        
        try:
            # Use DatabaseManager's get_stats method
            raw_stats = self.db.get_stats(days=30)
            
            for row in raw_stats:
                stats.append({
                    'date': row.get('date', ''),
                    'total_trades': row.get('total_trades', 0) or 0,
                    'successful_trades': row.get('successful_trades', 0) or 0,
                    'win_rate': row.get('win_rate', 0.0) or 0.0,
                    'realized_pnl': row.get('realized_pnl', 0.0) or 0.0
                })
        
        except Exception as e:
            logger.error(f"Error fetching daily stats: {e}")
        
        return stats
    
    def _filter_by_date(
        self,
        trades: List[TradeRecord],
        start_date: Optional[str],
        end_date: Optional[str]
    ) -> List[TradeRecord]:
        """Filter trades by date range"""
        filtered = []
        
        for trade in trades:
            try:
                trade_date = trade.timestamp[:10]  # YYYY-MM-DD
                
                if start_date and trade_date < start_date:
                    continue
                if end_date and trade_date > end_date:
                    continue
                
                filtered.append(trade)
            
            except Exception as e:
                logger.debug(f"Date filter error: {e}")
                continue
        
        return filtered


def export_trades_to_excel(db_manager, filepath: str):
    """
    Export trades to Excel format (requires openpyxl).
    
    Args:
        db_manager: DatabaseManager instance
        filepath: Output file path
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        
        exporter = TradeExporter(db_manager)
        trades = exporter.get_all_trades()
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Trades"
        
        # Header
        headers = list(asdict(trades[0]).keys()) if trades else []
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", fill_type="solid")
        
        # Data
        for row, trade in enumerate(trades, 2):
            for col, value in enumerate(asdict(trade).values(), 1):
                ws.cell(row=row, column=col, value=value)
        
        wb.save(filepath)
        logger.info(f"Exported {len(trades)} trades to Excel: {filepath}")
    
    except ImportError:
        logger.warning("openpyxl not installed, cannot export to Excel")
        raise
